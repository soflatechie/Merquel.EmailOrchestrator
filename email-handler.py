
#------------------------------------------------------------------------------
# Imports
#------------------------------------------------------------------------------
from dotenv import load_dotenv
from agents import Agent, Runner, trace, function_tool
from pydantic import BaseModel
from typing import Dict, List
import smtplib
import imaplib
import email
import os
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import asyncio

load_dotenv(override=True)

#------------------------------------------------------------------------------
# Pydantic models for strict schema compatibility
#------------------------------------------------------------------------------
class CategorizedItem(BaseModel):
    index: int
    category: str
    reason: str

# Module-level store so fetched email metadata never has to pass through the LLM output
_email_store: List[Dict] = []

#------------------------------------------------------------------------------
# Function to retrieve emails from Gmail inbox
#------------------------------------------------------------------------------
@function_tool
def get_gmail_inbox() -> str:
    """Retrieve the most recent 200 received emails across all Gmail folders.
    Returns a compact JSON list with index, sender, subject, date, and body snippet."""
    from datetime import timedelta
    import json
    global _email_store
    _email_store = []
    count = 200
    gmail_user = os.environ.get('GMAIL_USER')
    gmail_password = os.environ.get('GMAIL_APP_PASSWORD')
    since_str = (date.today() - timedelta(days=1)).strftime('%d-%b-%Y')
    seen_ids = set()
    try:
        with imaplib.IMAP4_SSL('imap.gmail.com', 993) as mail:
            mail.login(gmail_user, gmail_password)
            mail.select('"[Gmail]/All Mail"')
            _, data = mail.search(None, f'SINCE {since_str} NOT FROM "{gmail_user}"')
            email_ids = data[0].split()
            latest_ids = email_ids[-count:] if len(email_ids) >= count else email_ids
            print(f"Found {len(email_ids)} emails in All Mail since yesterday; fetching {len(latest_ids)}")
            for eid in reversed(latest_ids):
                _, msg_data = mail.fetch(eid, '(RFC822)')
                raw = msg_data[0][1]
                msg = email.message_from_bytes(raw)
                msg_id = msg.get('Message-ID', '')
                if msg_id in seen_ids:
                    continue
                seen_ids.add(msg_id)
                subject = msg.get('Subject', '(no subject)')
                sender = msg.get('From', '')
                date_str = msg.get('Date', '')
                body = ''
                if msg.is_multipart():
                    for part in msg.walk():
                        ctype = part.get_content_type()
                        if ctype == 'text/plain' and not part.get('Content-Disposition'):
                            body = part.get_payload(decode=True).decode(errors='replace')
                            break
                else:
                    body = msg.get_payload(decode=True).decode(errors='replace')
                _email_store.append({
                    'sender': sender,
                    'subject': subject,
                    'date': date_str,
                    'message_id': msg_id,
                })
    except Exception as e:
        print(f"IMAP ERROR: {e}")
    # Return compact list for LLM: only index + fields needed to classify
    compact = [
        {'index': i, 'sender': e['sender'], 'subject': e['subject'], 'body_snippet': ''}
        for i, e in enumerate(_email_store)
    ]
    return json.dumps(compact)


#------------------------------------------------------------------------------
# Function to write the categorized email list to an Excel file
#------------------------------------------------------------------------------
@function_tool
def write_csv_report(items: List[CategorizedItem]) -> str:
    """Write categorized email results to an Excel file.
    Each item has index (into the fetched email list), category, and reason."""
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), f'inbox_report_{date.today().strftime("%Y-%m-%d")}.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inbox Report'
    headers = ['Date', 'Sender', 'Subject', 'Category', 'Reason']
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for item in items:
        stored = _email_store[item.index] if item.index < len(_email_store) else {}
        row_num = ws.max_row + 1
        ws.append([
            stored.get('date', ''),
            stored.get('sender', ''),
            stored.get('subject', ''),
            item.category,
            item.reason,
        ])
        # Make the Subject cell a clickable link that opens the email in Gmail
        msg_id = stored.get('message_id', '').strip('<>').strip()
        if msg_id:
            from urllib.parse import quote
            gmail_url = f'https://mail.google.com/mail/u/0/#search/rfc822msgid%3A{quote(msg_id)}'
            subject_cell = ws.cell(row=row_num, column=3)
            subject_cell.hyperlink = gmail_url
            subject_cell.font = Font(color='0563C1', underline='single')
    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    wb.save(output_path)
    print(f"Excel report written: {output_path} ({len(items)} rows)")
    return f'success:{output_path}'


async def main():

    #------------------------------------------------------------------------------
    # Set up the email manager agent
    #------------------------------------------------------------------------------
    email_manager = Agent(
        name="Email Manager",
        instructions="""You are an email inbox categorization manager.

Follow these steps exactly:
1. Call get_gmail_inbox (no arguments) to retrieve emails. It returns a JSON list where each item has an 'index', 'sender', and 'subject'.
2. Using your own judgment, classify EVERY email into exactly ONE category:
   - SPAM/SCAM  — unsolicited junk, phishing, or fraudulent email
   - IMPORTANT  — requires a reply or follow-up action
   - MARKETING  — promotional, newsletter, or subscription email
   - OTHER      — anything that does not fit the above
3. Call write_csv_report ONCE with the full list. Each entry must have:
   - index: the integer index from the email list
   - category: one of SPAM/SCAM, IMPORTANT, MARKETING, OTHER
   - reason: one short sentence
   You MUST include every single email — do not skip any.""",
        tools=[get_gmail_inbox, write_csv_report],
        model="gpt-4o-mini",
    )

    print('Email manager agent initialized')

    #------------------------------------------------------------------------------
    # Run the email manager
    #------------------------------------------------------------------------------
    with trace("Email Inbox Categorizer"):
        print("Retrieving and categorizing today's emails...")
        result = await Runner.run(
            email_manager,
            "Retrieve today's emails from the inbox and categorize each one.",
            max_turns=10,
        )
    print('\n' + '=' * 60)
    print('EMAIL CATEGORIZATION RESULTS')
    print('=' * 60)
    print(result.final_output)


if __name__ == "__main__":
    print('main executing')
    asyncio.run(main())
